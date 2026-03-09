"""
Indian Founder Personal Finance Planner — Excel Workbook Generator
Usage: python generate_excel.py [output_path]
Generates a fully-featured, formula-driven .xlsx financial planning workbook.
"""

import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = sys.argv[1] if len(sys.argv) > 1 else "founder_finance_plan_india.xlsx"

# ── Color Palette ─────────────────────────────────────────────────────────────
NAVY   = "1B3A6B"
TEAL   = "0E6B6B"
ORANGE = "C95E1E"
GREEN  = "1A6B3C"
RED_BG = "C0392B"
LGRAY  = "F2F4F7"
MGRAY  = "D5D8DC"
YELLOW = "FFF3CD"
WHITE  = "FFFFFF"
BLUE_INPUT = "1F4E79"   # blue text for user-editable inputs

# ── Style Helpers ─────────────────────────────────────────────────────────────
def ft(bold=False, size=11, color="000000", italic=False):
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border(sides="all"):
    s = Side(style="thin", color="BBBBBB")
    n = Side(style=None)
    if sides == "all":
        return Border(left=s, right=s, top=s, bottom=s)
    if sides == "bottom":
        return Border(bottom=s)
    if sides == "outer":
        thick = Side(style="medium", color="888888")
        return Border(left=thick, right=thick, top=thick, bottom=thick)
    return Border()

def header_row(ws, row, cols, texts, bg=NAVY, fg=WHITE, size=11, bold=True, height=22):
    ws.row_dimensions[row].height = height
    for col, text in zip(cols, texts):
        c = ws.cell(row=row, column=col, value=text)
        c.font = ft(bold=bold, size=size, color=fg)
        c.fill = fill(bg)
        c.alignment = align("center")
        c.border = thin_border()

def section_title(ws, row, col, text, span_end_col, bg=TEAL):
    ws.row_dimensions[row].height = 20
    c = ws.cell(row=row, column=col, value=text)
    c.font = ft(bold=True, size=12, color=WHITE)
    c.fill = fill(bg)
    c.alignment = align("left")
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=span_end_col)

def label(ws, row, col, text, bold=False, indent=False):
    v = ("    " + text) if indent else text
    c = ws.cell(row=row, column=col, value=v)
    c.font = ft(bold=bold, size=10)
    c.alignment = align("left", wrap=True)
    c.border = thin_border("bottom")
    return c

def input_cell(ws, row, col, value, fmt="inr", note=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = ft(bold=False, size=10, color=BLUE_INPUT)
    c.fill = fill(YELLOW)
    c.alignment = align("right")
    c.border = thin_border()
    if fmt == "inr":
        c.number_format = u'[\u20B9]#,##0;([\u20B9]#,##0);"-"'
    elif fmt == "pct":
        c.number_format = "0.0%"
    elif fmt == "num":
        c.number_format = "#,##0"
    elif fmt == "text":
        c.number_format = "@"
    if note:
        from openpyxl.comments import Comment
        c.comment = Comment(note, "Skill")
    return c

def formula_cell(ws, row, col, formula, fmt="inr", color="000000", bg=None):
    c = ws.cell(row=row, column=col, value=formula)
    c.font = ft(size=10, color=color)
    if bg:
        c.fill = fill(bg)
    c.alignment = align("right")
    c.border = thin_border()
    if fmt == "inr":
        c.number_format = u'[\u20B9]#,##0;([\u20B9]#,##0);"-"'
    elif fmt == "pct":
        c.number_format = "0.0%"
    elif fmt == "num":
        c.number_format = "#,##0.0"
    elif fmt == "months":
        c.number_format = '0.0" mo"'
    return c

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

# ═════════════════════════════════════════════════════════════════════════════
wb = Workbook()

# ── SHEET 1: DASHBOARD ───────────────────────────────────────────────────────
dash = wb.active
dash.title = "📊 Dashboard"
dash.sheet_view.showGridLines = False
set_col_widths(dash, {"A": 3, "B": 28, "C": 18, "D": 18, "E": 18, "F": 18, "G": 3})

# Title banner
dash.row_dimensions[1].height = 10
dash.row_dimensions[2].height = 36
dash.merge_cells("B2:F2")
t = dash["B2"]
t.value = "🇮🇳  Indian Founder Personal Finance Planner"
t.font = ft(bold=True, size=18, color=WHITE)
t.fill = fill(NAVY)
t.alignment = align("center", "center")

dash.row_dimensions[3].height = 18
dash.merge_cells("B3:F3")
sub = dash["B3"]
sub.value = "Update your inputs on the '⚙️ Inputs' sheet — all numbers here update automatically"
sub.font = ft(italic=True, size=10, color="555555")
sub.alignment = align("center")

# ── Section: Financial Health Scorecard
dash.row_dimensions[5].height = 8
section_title(dash, 6, 2, "  📋  FINANCIAL HEALTH SCORECARD", 6)

metrics = [
    ("Monthly Take-home (₹)",     "='⚙️ Inputs'!C7"),
    ("Monthly Burn (₹)",          "='⚙️ Inputs'!C14"),
    ("Monthly Surplus / Deficit",  "='⚙️ Inputs'!C7-'⚙️ Inputs'!C14"),
    ("Savings Rate",              "=IF('⚙️ Inputs'!C7>0,('⚙️ Inputs'!C7-'⚙️ Inputs'!C14)/'⚙️ Inputs'!C7,0)"),
    ("Emergency Fund (months)",   "=IF('⚙️ Inputs'!C14>0,'⚙️ Inputs'!C20/'⚙️ Inputs'!C14,0)"),
    ("Personal Runway (months)",  "=IF(('⚙️ Inputs'!C14-'⚙️ Inputs'!C7)>0,'⚙️ Inputs'!C20/('⚙️ Inputs'!C14-'⚙️ Inputs'!C7),99)"),
    ("EMI-to-Income Ratio",       "=IF('⚙️ Inputs'!C7>0,'⚙️ Inputs'!C17/'⚙️ Inputs'!C7,0)"),
    ("Annual Tax Saving Available","='💸 Tax Planner'!C38"),
]
fmts = ["inr","inr","inr","pct","months","months","pct","inr"]
statuses = [
    ("", ""),
    ("", ""),
    ("✅ Surplus" ,"🔴 Deficit"),
    ("✅ ≥20%","🔴 <10%"),
    ("✅ ≥6 mo","🔴 <3 mo"),
    ("✅ ≥12 mo","🔴 <6 mo"),
    ("✅ <25%","🔴 >40%"),
    ("", ""),
]

header_row(dash, 7, [2,3,4,5], ["Metric","Value","Status","Benchmark"], bg=TEAL)
for i, (metric, formula) in enumerate(metrics):
    r = 8 + i
    dash.row_dimensions[r].height = 18
    label(dash, r, 2, metric)
    formula_cell(dash, r, 3, formula, fmt=fmts[i])
    # Status formula using IF
    ok, bad = statuses[i]
    if ok:
        formula_cell(dash, r, 4,
            f'=IF({formula[1:]}>=0,"{ok}","{bad}")',
            fmt="text", color=GREEN)
    else:
        dash.cell(row=r, column=4).value = "—"
        dash.cell(row=r, column=4).font = ft(size=10)
    bench_vals = ["","","","≥20%","≥6 mo","≥12 mo","<25%",""]
    dash.cell(row=r, column=5, value=bench_vals[i]).font = ft(size=10, italic=True, color="555555")
    dash.cell(row=r, column=5).alignment = align("center")

# ── Section: Investment Stack Status
r = 17
dash.row_dimensions[r].height = 8
section_title(dash, r+1, 2, "  💼  YOUR INVESTMENT STACK STATUS", 6, bg=TEAL)
header_row(dash, r+2, [2,3,4,5,6], ["Layer","Instrument","Target/yr","Invested/yr","Gap"], bg=TEAL, height=18)

stack = [
    ("🔴 Survival", "Emergency Fund (6× burn)", "='⚙️ Inputs'!C14*6", "='⚙️ Inputs'!C20", "=D{r}-E{r}"),
    ("🟡 Tax Shield","PPF (Section 80C)",         "=150000",            "='💼 Investments'!C8", "=D{r}-E{r}"),
    ("🟡 Tax Shield","NPS Tier 1 (80CCD 1B)",     "=50000",             "='💼 Investments'!C9", "=D{r}-E{r}"),
    ("🟡 Tax Shield","Health Ins Premium (80D)",   "=25000",             "='⚙️ Inputs'!C32",     "=D{r}-E{r}"),
    ("🟢 Wealth",   "Index Fund SIP (Nifty 50)",  "='💼 Investments'!C12*12","='💼 Investments'!C12*12","=0"),
    ("🟢 Wealth",   "Flexicap SIP",               "='💼 Investments'!C13*12","='💼 Investments'!C13*12","=0"),
]
for j, (layer, inst, tgt, invested_formula, gap) in enumerate(stack):
    rr = r + 3 + j
    dash.row_dimensions[rr].height = 17
    dash.cell(row=rr, column=2, value=layer).font = ft(size=10, bold=True)
    dash.cell(row=rr, column=3, value=inst).font = ft(size=10)
    formula_cell(dash, rr, 4, tgt)
    formula_cell(dash, rr, 5, inv, color="006400")
    formula_cell(dash, rr, 6, gap.replace("{r}", str(rr)), color=RED_BG)
    for col in [2,3,4,5,6]:
        dash.cell(row=rr, column=col).border = thin_border("bottom")
        dash.cell(row=rr, column=col).fill = fill(LGRAY if j%2==0 else WHITE)
        if col in [4,5,6]:
            dash.cell(row=rr, column=col).number_format = u'[\u20B9]#,##0;([\u20B9]#,##0);"-"'

# ── Section: Quick Alerts
r2 = r + 10
section_title(dash, r2, 2, "  ⚡  QUICK ACTION ALERTS", 6, bg=ORANGE)
alerts = [
    "🔴 Emergency fund < 3 months → Top priority before any investing",
    "🟡 PPF not maxed → Fill ₹1.5L before 5 April every year for EEE benefit",
    "🟡 NPS 80CCD(1B) unused → ₹50,000 saves ₹15,600 in tax (30% bracket)",
    "🔴 No personal health insurance → Buy before a medical event, not after",
    "🟡 Advance tax due → 15 Jun / 15 Sep / 15 Dec / 15 Mar (see Tax sheet)",
]
for k, alert in enumerate(alerts):
    rr = r2 + 1 + k
    dash.row_dimensions[rr].height = 17
    c = dash.cell(row=rr, column=2, value=alert)
    c.font = ft(size=10)
    c.fill = fill(LGRAY if k%2==0 else WHITE)
    dash.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=6)
    c.alignment = align("left")

# ═════════════════════════════════════════════════════════════════════════════
# ── SHEET 2: INPUTS ──────────────────────────────────────────────────────────
inp = wb.create_sheet("⚙️ Inputs")
inp.sheet_view.showGridLines = False
set_col_widths(inp, {"A": 3, "B": 32, "C": 18, "D": 30, "E": 3})

inp.row_dimensions[1].height = 10
inp.row_dimensions[2].height = 32
inp.merge_cells("B2:D2")
t2 = inp["B2"]
t2.value = "⚙️  Your Profile & Inputs  [Edit BLUE cells only]"
t2.font = ft(bold=True, size=15, color=WHITE)
t2.fill = fill(NAVY)
t2.alignment = align("center", "center")

inp.row_dimensions[3].height = 16
inp.merge_cells("B3:D3")
inp["B3"].value = "Blue cells = your inputs | Yellow background = key inputs | All other sheets update automatically"
inp["B3"].font = ft(italic=True, size=9, color="555555")
inp["B3"].alignment = align("center")

def inp_row(ws, row, lbl, val, fmt="inr", note=None, bold_lbl=False):
    ws.row_dimensions[row].height = 18
    label(ws, row, 2, lbl, bold=bold_lbl)
    return input_cell(ws, row, 3, val, fmt=fmt, note=note)

# Personal profile
section_title(inp, 5, 2, "  👤  PERSONAL PROFILE", 4)
inp_row(inp, 6,  "Your Name",               "Founder Name",  fmt="text")
inp_row(inp, 7,  "Monthly Take-home (₹)",   80000,  note="After TDS deducted from startup salary. Enter 0 if bootstrapped.")
inp_row(inp, 8,  "Age (years)",             30,     fmt="num")
inp_row(inp, 9,  "City",                    "Bengaluru", fmt="text")
inp_row(inp, 10, "Dependents",              "Spouse, 1 child", fmt="text")

# Monthly expenses
section_title(inp, 12, 2, "  🏠  MONTHLY EXPENSES (₹)", 4)
inp_row(inp, 13, "Rent / Home Loan EMI",    25000)
inp_row(inp, 14, "Food & Groceries",        8000)
inp_row(inp, 15, "Transport",               4000)
inp_row(inp, 16, "Subscriptions & Tools",   3000)
inp_row(inp, 17, "Other EMIs (loans etc.)", 0,    note="Car loan, education loan, personal loan EMIs")
inp_row(inp, 18, "Family / Parent Support", 5000)
inp_row(inp, 19, "Miscellaneous",           5000)
# Total burn formula
label(inp, 20, 2, "Total Monthly Burn (₹)", bold=True)
formula_cell(inp, 20, 3, "=SUM(C13:C19)", bg=LGRAY)

# Savings & investments
section_title(inp, 22, 2, "  💰  SAVINGS & EXISTING INVESTMENTS (₹)", 4)
inp_row(inp, 23, "Cash in Savings Account", 200000)
inp_row(inp, 24, "Fixed Deposits (FDs)",    50000)
inp_row(inp, 25, "Liquid Mutual Funds",     0)
# Total savings
label(inp, 26, 2, "Total Liquid Savings", bold=True)
formula_cell(inp, 26, 3, "=SUM(C23:C25)", bg=LGRAY)

inp_row(inp, 27, "PPF Balance",             80000)
inp_row(inp, 28, "NPS Balance",             0)
inp_row(inp, 29, "Mutual Fund Portfolio",   50000)
inp_row(inp, 30, "EPF (from prior job)",    0)
# Total net worth
label(inp, 31, 2, "Total Investments (ex-equity)", bold=True)
formula_cell(inp, 31, 3, "=SUM(C27:C30)", bg=LGRAY)

# Insurance
section_title(inp, 33, 2, "  🛡️  INSURANCE", 4)
inp_row(inp, 34, "Health Insurance Cover (₹)", 0,     note="0 if you have no personal policy")
inp_row(inp, 35, "Term Life Cover (₹)",        0)
inp_row(inp, 36, "Health Premium Paid/yr (₹)", 0,     note="For 80D deduction calculation")

# Startup & equity
section_title(inp, 38, 2, "  🚀  STARTUP & EQUITY", 4)
inp_row(inp, 39, "Startup Stage",            "Bootstrapped", fmt="text")
inp_row(inp, 40, "Your Equity Stake (%)",    0.5,  fmt="pct")
inp_row(inp, 41, "ESOPs Granted (shares)",   0,    fmt="num")
inp_row(inp, 42, "ESOP Strike Price (₹)",    0)
inp_row(inp, 43, "ESOP FMV Today (₹)",       0)

# Tax
section_title(inp, 45, 2, "  🧾  TAX DETAILS", 4)
inp_row(inp, 46, "Tax Regime",               "Old", fmt="text", note="Enter: Old or New")
inp_row(inp, 47, "HRA Received / yr (₹)",   0,     note="Only if on salary; 0 if no HRA component")
inp_row(inp, 48, "Home Loan Interest / yr",  0)
inp_row(inp, 49, "LTA Claimed / yr (₹)",     0)

# ═════════════════════════════════════════════════════════════════════════════
# ── SHEET 3: CASH FLOW ───────────────────────────────────────────────────────
cf = wb.create_sheet("💵 Cash Flow")
cf.sheet_view.showGridLines = False
set_col_widths(cf, {"A": 3, "B": 30, "C": 16, "D": 16, "E": 16, "F": 16, "G": 16, "H": 16, "I": 3})

cf.row_dimensions[2].height = 32
cf.merge_cells("B2:H2")
th = cf["B2"]
th.value = "💵  Monthly Cash Flow Planner  (12-Month View)"
th.font = ft(bold=True, size=15, color=WHITE)
th.fill = fill(NAVY)
th.alignment = align("center", "center")

months = ["Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec","Jan","Feb","Mar"]
header_row(cf, 4, range(2, 15), ["Category"] + months, bg=TEAL, size=10)

cf_items = [
    ("INCOME", None, True),
    ("Take-home Salary", "='⚙️ Inputs'!C7", False),
    ("Other Income",     0, False),
    ("Total Income",     None, True),
    ("", None, False),
    ("EXPENSES", None, True),
    ("Rent / Home EMI",  "='⚙️ Inputs'!C13", False),
    ("Food & Groceries", "='⚙️ Inputs'!C14", False),
    ("Transport",        "='⚙️ Inputs'!C15", False),
    ("Subscriptions",    "='⚙️ Inputs'!C16", False),
    ("Other EMIs",       "='⚙️ Inputs'!C17", False),
    ("Family Support",   "='⚙️ Inputs'!C18", False),
    ("Miscellaneous",    "='⚙️ Inputs'!C19", False),
    ("Total Expenses",   None, True),
    ("", None, False),
    ("NET SURPLUS / (DEFICIT)", None, True),
    ("Cumulative Cash Position", None, True),
]

# Map row numbers
section_rows = {
    "INCOME": 5, "Total Income": 8,
    "EXPENSES": 10, "Total Expenses": 19,
    "NET SURPLUS / (DEFICIT)": 21,
    "Cumulative Cash Position": 22,
}
row_map = {}
r = 5
for name, default, is_section in cf_items:
    cf.row_dimensions[r].height = 17
    if not name:
        r += 1
        continue
    row_map[name] = r
    if is_section:
        c = cf.cell(row=r, column=2, value=name)
        c.font = ft(bold=True, size=10, color=WHITE)
        c.fill = fill(TEAL if "INCOME" in name or "EXPENSES" in name else NAVY)
        cf.merge_cells(start_row=r, start_column=2, end_row=r, end_column=2)
    else:
        label(cf, r, 2, "  " + name)
    for m_idx in range(12):
        col = 3 + m_idx
        if is_section:
            cf.cell(row=r, column=col).fill = fill(TEAL if "INCOME" in name or "EXPENSES" in name else NAVY)
        elif default is not None:
            if isinstance(default, str):
                formula_cell(cf, r, col, default, color=GREEN)
            else:
                input_cell(cf, r, col, default)
    r += 1

# Total Income formula
ti_r = row_map["Total Income"]
for col in range(3, 15):
    formula_cell(cf, ti_r, col, f"=C{row_map['Take-home Salary']}+C{row_map['Other Income']}", bg=LGRAY)
    cf.cell(row=ti_r, column=col).column = col

# Adjust for month columns properly
for m_idx in range(12):
    col = 3 + m_idx
    col_l = get_column_letter(col)
    # Total Income
    cf.cell(row=ti_r, column=col).value = f"=SUM({col_l}{row_map['Take-home Salary']}:{col_l}{row_map['Other Income']})"
    cf.cell(row=ti_r, column=col).number_format = u'[\u20B9]#,##0;([\u20B9]#,##0);"-"'
    cf.cell(row=ti_r, column=col).fill = fill(LGRAY)
    cf.cell(row=ti_r, column=col).font = ft(bold=True, size=10)

    # Total Expenses
    te_r = row_map["Total Expenses"]
    exp_start = row_map["Rent / Home EMI"]
    exp_end = row_map["Miscellaneous"]
    cf.cell(row=te_r, column=col).value = f"=SUM({col_l}{exp_start}:{col_l}{exp_end})"
    cf.cell(row=te_r, column=col).number_format = u'[\u20B9]#,##0;([\u20B9]#,##0);"-"'
    cf.cell(row=te_r, column=col).fill = fill(LGRAY)
    cf.cell(row=te_r, column=col).font = ft(bold=True, size=10)

    # Net Surplus
    ns_r = row_map["NET SURPLUS / (DEFICIT)"]
    cf.cell(row=ns_r, column=col).value = f"={col_l}{ti_r}-{col_l}{te_r}"
    cf.cell(row=ns_r, column=col).number_format = u'[\u20B9]#,##0;([\u20B9]#,##0);"-"'
    cf.cell(row=ns_r, column=col).font = ft(bold=True, size=10, color=BLUE_INPUT)
    cf.cell(row=ns_r, column=col).fill = fill(YELLOW)

    # Cumulative
    cum_r = row_map["Cumulative Cash Position"]
    if m_idx == 0:
        cf.cell(row=cum_r, column=col).value = f"='⚙️ Inputs'!C26+{col_l}{ns_r}"
    else:
        prev_col = get_column_letter(col - 1)
        cf.cell(row=cum_r, column=col).value = f"={prev_col}{cum_r}+{col_l}{ns_r}"
    cf.cell(row=cum_r, column=col).number_format = u'[\u20B9]#,##0;([\u20B9]#,##0);"-"'
    cf.cell(row=cum_r, column=col).font = ft(bold=True, size=10, color=GREEN)
    cf.cell(row=cum_r, column=col).fill = fill(LGRAY)

# Runway & Savings Rate KPIs
kpi_r = row_map["Cumulative Cash Position"] + 2
section_title(cf, kpi_r, 2, "  📊  KEY METRICS", 6)
kpi_items = [
    ("Personal Runway (months)", "=IF(('⚙️ Inputs'!C20-'⚙️ Inputs'!C7)>0,'⚙️ Inputs'!C26/('⚙️ Inputs'!C20-'⚙️ Inputs'!C7),99)", "months"),
    ("Monthly Savings Rate",      "=IF('⚙️ Inputs'!C7>0,('⚙️ Inputs'!C7-'⚙️ Inputs'!C20)/'⚙️ Inputs'!C7,0)", "pct"),
    ("Emergency Fund Target (₹)", "='⚙️ Inputs'!C20*6", "inr"),
    ("Emergency Fund Current (₹)","='⚙️ Inputs'!C26", "inr"),
    ("Emergency Fund Gap (₹)",    "=C{r2}-C{r3}".replace("{r2}", str(kpi_r+3)).replace("{r3}", str(kpi_r+4)), "inr"),
    ("EMI-to-Income Ratio",       "=IF('⚙️ Inputs'!C7>0,'⚙️ Inputs'!C17/'⚙️ Inputs'!C7,0)", "pct"),
]
for j, (lbl, frm, fmt) in enumerate(kpi_items):
    rr = kpi_r + 1 + j
    cf.row_dimensions[rr].height = 18
    label(cf, rr, 2, lbl)
    formula_cell(cf, rr, 3, frm, fmt=fmt)

# ═════════════════════════════════════════════════════════════════════════════
# ── SHEET 4: TAX PLANNER ─────────────────────────────────────────────────────
tax = wb.create_sheet("💸 Tax Planner")
tax.sheet_view.showGridLines = False
set_col_widths(tax, {"A": 3, "B": 34, "C": 20, "D": 20, "E": 20, "F": 3})

tax.row_dimensions[2].height = 32
tax.merge_cells("B2:E2")
th3 = tax["B2"]
th3.value = "💸  Tax Planner  —  Old vs New Regime Comparison (FY 2024-25)"
th3.font = ft(bold=True, size=14, color=WHITE)
th3.fill = fill(NAVY)
th3.alignment = align("center", "center")

header_row(tax, 4, [2,3,4,5], ["Component","Old Regime","New Regime","Difference (Old-New)"], bg=TEAL)

# Gross income
label(tax, 5, 2, "Gross Annual Income (Salary)", bold=True)
formula_cell(tax, 5, 3, "='⚙️ Inputs'!C7*12")
formula_cell(tax, 5, 4, "='⚙️ Inputs'!C7*12")
formula_cell(tax, 5, 5, "=C5-D5")

# Deductions — Old Regime
section_title(tax, 7, 2, "  DEDUCTIONS (Old Regime Only)", 4, bg=MGRAY)
old_deductions = [
    ("Standard Deduction (Sec 16)",      "=50000"),
    ("HRA Exemption",                     "='⚙️ Inputs'!C47"),
    ("80C: PPF + ELSS + LIC + EPF",       "=MIN('💼 Investments'!C8+'💼 Investments'!C9,150000)"),
    ("80CCD(1B): NPS Tier 1",             "=MIN('💼 Investments'!C10,50000)"),
    ("80D: Health Ins Premium",           "=MIN('⚙️ Inputs'!C36,25000)"),
    ("Home Loan Interest (Sec 24b)",      "=MIN('⚙️ Inputs'!C48,200000)"),
    ("LTA",                               "='⚙️ Inputs'!C49"),
]
ded_start = 8
for j, (name, formula) in enumerate(old_deductions):
    rr = ded_start + j
    tax.row_dimensions[rr].height = 17
    label(tax, rr, 2, "  " + name)
    formula_cell(tax, rr, 3, formula)
    tax.cell(row=rr, column=4).value = "—"
    tax.cell(row=rr, column=4).font = ft(size=10, color="888888")
    tax.cell(row=rr, column=4).alignment = align("center")
    tax.cell(row=rr, column=4).border = thin_border("bottom")

# Total deductions
td_r = ded_start + len(old_deductions)
label(tax, td_r, 2, "Total Deductions (Old Regime)", bold=True)
formula_cell(tax, td_r, 3, f"=SUM(C{ded_start}:C{td_r-1})", bg=LGRAY)
tax.cell(row=td_r, column=4, value=75000).number_format = u'[\u20B9]#,##0;([\u20B9]#,##0);"-"'
tax.cell(row=td_r, column=4).font = ft(size=10, italic=True, color="555555")
tax.cell(row=td_r, column=4).border = thin_border()
tax.cell(row=td_r, column=4).alignment = align("right")

# Taxable income
ti_tax_r = td_r + 2
label(tax, ti_tax_r, 2, "Taxable Income", bold=True)
formula_cell(tax, ti_tax_r, 3, f"=MAX(C5-C{td_r},0)", bg=YELLOW)
formula_cell(tax, ti_tax_r, 4, "=MAX(C5-75000,0)", bg=YELLOW)
formula_cell(tax, ti_tax_r, 5, f"=C{ti_tax_r}-D{ti_tax_r}")

# Tax calculation (Old Regime slabs FY24-25)
tax_r = ti_tax_r + 2
section_title(tax, tax_r, 2, "  TAX CALCULATION (FY 2024-25 Slabs)", 4, bg=MGRAY)

old_tax_formula = (
    f"=MAX(0,"
    f"IF(C{ti_tax_r}<=250000,0,"
    f"IF(C{ti_tax_r}<=500000,(C{ti_tax_r}-250000)*0.05,"
    f"IF(C{ti_tax_r}<=1000000,12500+(C{ti_tax_r}-500000)*0.2,"
    f"112500+(C{ti_tax_r}-1000000)*0.3))))"
)
new_tax_formula = (
    f"=MAX(0,"
    f"IF(D{ti_tax_r}<=300000,0,"
    f"IF(D{ti_tax_r}<=700000,(D{ti_tax_r}-300000)*0.05,"
    f"IF(D{ti_tax_r}<=1000000,20000+(D{ti_tax_r}-700000)*0.1,"
    f"IF(D{ti_tax_r}<=1200000,50000+(D{ti_tax_r}-1000000)*0.15,"
    f"IF(D{ti_tax_r}<=1500000,80000+(D{ti_tax_r}-1200000)*0.2,"
    f"140000+(D{ti_tax_r}-1500000)*0.3))))))"
)

label(tax, tax_r+1, 2, "  Income Tax (before cess)")
formula_cell(tax, tax_r+1, 3, old_tax_formula)
formula_cell(tax, tax_r+1, 4, new_tax_formula)
formula_cell(tax, tax_r+1, 5, f"=C{tax_r+1}-D{tax_r+1}")

label(tax, tax_r+2, 2, "  Rebate u/s 87A")
formula_cell(tax, tax_r+2, 3, f"=IF(C{ti_tax_r}<=500000,MIN(C{tax_r+1},12500),0)")
formula_cell(tax, tax_r+2, 4, f"=IF(D{ti_tax_r}<=700000,MIN(D{tax_r+1},25000),0)")
formula_cell(tax, tax_r+2, 5, f"=C{tax_r+2}-D{tax_r+2}")

label(tax, tax_r+3, 2, "  Health & Education Cess (4%)")
formula_cell(tax, tax_r+3, 3, f"=(C{tax_r+1}-C{tax_r+2})*0.04")
formula_cell(tax, tax_r+3, 4, f"=(D{tax_r+1}-D{tax_r+2})*0.04")
formula_cell(tax, tax_r+3, 5, f"=C{tax_r+3}-D{tax_r+3}")

total_tax_r = tax_r + 5
label(tax, total_tax_r, 2, "TOTAL TAX PAYABLE", bold=True)
formula_cell(tax, total_tax_r, 3, f"=(C{tax_r+1}-C{tax_r+2})+C{tax_r+3}", bg=YELLOW)
formula_cell(tax, total_tax_r, 4, f"=(D{tax_r+1}-D{tax_r+2})+D{tax_r+3}", bg=YELLOW)
formula_cell(tax, total_tax_r, 5, f"=C{total_tax_r}-D{total_tax_r}")

verdict_r = total_tax_r + 2
tax.row_dimensions[verdict_r].height = 22
tax.merge_cells(f"B{verdict_r}:E{verdict_r}")
vf = tax.cell(row=verdict_r, column=2)
vf.value = f'=IF(C{total_tax_r}<D{total_tax_r},"✅ VERDICT: Old Regime saves ₹"&TEXT(D{total_tax_r}-C{total_tax_r},"#,##0")&" vs New Regime → Stick with OLD","✅ VERDICT: New Regime saves ₹"&TEXT(C{total_tax_r}-D{total_tax_r},"#,##0")&" vs Old Regime → Switch to NEW")'
vf.font = ft(bold=True, size=12, color=WHITE)
vf.fill = fill(GREEN)
vf.alignment = align("center", "center")

# Tax saving opportunities
opp_r = verdict_r + 2
section_title(tax, opp_r, 2, "  💡  UNUSED TAX-SAVING OPPORTUNITIES", 4, bg=ORANGE)
opportunities = [
    ("NPS 80CCD(1B) unused room",   f"=MAX(0,50000-MIN('💼 Investments'!C10,50000))"),
    ("80C unused room",             f"=MAX(0,150000-MIN('💼 Investments'!C8+'💼 Investments'!C9,150000))"),
    ("80D unused room (self)",       f"=MAX(0,25000-MIN('⚙️ Inputs'!C36,25000))"),
    ("Total potential tax saving",   f"=SUM(C{opp_r+1}:C{opp_r+3})*0.3"),
]
for j, (lbl, frm) in enumerate(opportunities):
    rr = opp_r + 1 + j
    tax.row_dimensions[rr].height = 17
    label(tax, rr, 2, "  " + lbl, bold=(j==3))
    formula_cell(tax, rr, 3, frm, bg=LGRAY if j==3 else WHITE)

# Store total tax saving for dashboard reference
tax.cell(row=38, column=3).value = f"=C{opp_r+4}"
tax.cell(row=38, column=3).number_format = u'[\u20B9]#,##0;([\u20B9]#,##0);"-"'

# Advance Tax Calendar
adv_r = opp_r + 7
section_title(tax, adv_r, 2, "  📅  ADVANCE TAX CALENDAR (Mark in your diary!)", 4, bg=TEAL)
header_row(tax, adv_r+1, [2,3,4,5], ["Instalment","Due Date","% of Annual Tax","Estimated Amount"], bg=TEAL)
adv_data = [
    ("1st Instalment", "15 June 2025",      "15%", 0.15),
    ("2nd Instalment", "15 September 2025", "45%", 0.45),
    ("3rd Instalment", "15 December 2025",  "75%", 0.75),
    ("4th Instalment", "15 March 2026",     "100%",1.00),
]
for j, (inst, dt, pct, mul) in enumerate(adv_data):
    rr = adv_r + 2 + j
    tax.row_dimensions[rr].height = 17
    tax.cell(row=rr, column=2, value=inst).font = ft(size=10)
    tax.cell(row=rr, column=3, value=dt).font = ft(size=10)
    tax.cell(row=rr, column=4, value=pct).font = ft(size=10)
    tax.cell(row=rr, column=4).alignment = align("center")
    formula_cell(tax, rr, 5, f"=C{total_tax_r}*{mul}")
    for col in [2,3,4,5]:
        tax.cell(row=rr, column=col).border = thin_border("bottom")
        tax.cell(row=rr, column=col).fill = fill(LGRAY if j%2==0 else WHITE)

# ═════════════════════════════════════════════════════════════════════════════
# ── SHEET 5: INVESTMENTS ─────────────────────────────────────────────────────
inv = wb.create_sheet("💼 Investments")
inv.sheet_view.showGridLines = False
set_col_widths(inv, {"A": 3, "B": 30, "C": 18, "D": 18, "E": 18, "F": 20, "G": 3})

inv.row_dimensions[2].height = 32
inv.merge_cells("B2:F2")
th4 = inv["B2"]
th4.value = "💼  Investment Tracker & Planner"
th4.font = ft(bold=True, size=15, color=WHITE)
th4.fill = fill(NAVY)
th4.alignment = align("center", "center")

header_row(inv, 4, [2,3,4,5,6], ["Instrument","Monthly SIP / Contrib","Annual Amount","Current Balance","Target / Year"], bg=TEAL)

# Layer 1 - Emergency
section_title(inv, 5, 2, "  🔴  LAYER 1 — SURVIVAL (Emergency Fund)", 6)
label(inv, 6, 2, "  Liquid Mutual Fund (Emergency)")
input_cell(inv, 6, 3, 0, note="Monthly top-up until emergency fund is complete")
formula_cell(inv, 6, 4, "=C6*12")
formula_cell(inv, 6, 5, "='⚙️ Inputs'!C25", color=GREEN)
formula_cell(inv, 6, 6, "='⚙️ Inputs'!C20*6")   # 6x burn target

# Layer 2 - Tax shield
section_title(inv, 7, 2, "  🟡  LAYER 2 — TAX SHIELD", 6)
tax_instruments = [
    ("  PPF (Section 80C)",         0,  150000,  "='⚙️ Inputs'!C27"),
    ("  NPS Tier 1 (80CCD 1B)",     0,  50000,   "='⚙️ Inputs'!C28"),
    ("  ELSS SIP (80C remainder)",  0,  0,        ""),
    ("  Health Insurance Premium",  0,  25000,   "='⚙️ Inputs'!C36"),
]
for j, (name, monthly, annual_tgt, bal_formula) in enumerate(tax_instruments):
    rr = 8 + j
    inv.row_dimensions[rr].height = 17
    label(inv, rr, 2, name)
    input_cell(inv, rr, 3, monthly, note="Monthly contribution")
    formula_cell(inv, rr, 4, f"=C{rr}*12")
    if bal_formula:
        formula_cell(inv, rr, 5, bal_formula, color=GREEN)
    else:
        input_cell(inv, rr, 5, 0)
    inv.cell(row=rr, column=6, value=annual_tgt).number_format = u'[\u20B9]#,##0;([\u20B9]#,##0);"-"'
    inv.cell(row=rr, column=6).font = ft(size=10, color="555555", italic=True)
    inv.cell(row=rr, column=6).border = thin_border("bottom")
    inv.cell(row=rr, column=6).alignment = align("right")

# Layer 3 - Wealth
section_title(inv, 12, 2, "  🟢  LAYER 3 — WEALTH BUILDING", 6)
wealth_instruments = [
    ("  Nifty 50 Index Fund SIP",  0, "Direct plan | UTI / Nippon | 0.1% expense"),
    ("  Flexicap Fund SIP",        0, "Direct plan | Parag Parikh | 0.6% expense"),
    ("  NPS Tier 2 (no lock-in)",  0, "Flexible; equity + debt mix"),
]
for j, (name, monthly, note) in enumerate(wealth_instruments):
    rr = 13 + j
    inv.row_dimensions[rr].height = 17
    label(inv, rr, 2, name)
    input_cell(inv, rr, 3, monthly, note=note)
    formula_cell(inv, rr, 4, f"=C{rr}*12")
    input_cell(inv, rr, 5, 0)
    inv.cell(row=rr, column=6, value=note).font = ft(size=9, italic=True, color="777777")
    inv.cell(row=rr, column=6).alignment = align("left", wrap=True)
    inv.cell(row=rr, column=6).border = thin_border("bottom")

# Totals
inv.row_dimensions[17].height = 18
label(inv, 17, 2, "Total Monthly Investment", bold=True)
formula_cell(inv, 17, 3, "=SUM(C6:C15)", bg=LGRAY)
formula_cell(inv, 17, 4, "=SUM(D6:D15)", bg=LGRAY)

# ── Compounding Projection Table
proj_r = 19
section_title(inv, proj_r, 2, "  📈  SIP COMPOUNDING PROJECTION", 6, bg=TEAL)
header_row(inv, proj_r+1, [2,3,4,5,6], ["Year","Monthly SIP","CAGR Assumed","Corpus (Conservative 10%)","Corpus (Optimistic 13%)"], bg=TEAL, size=10)
for yr in range(1, 11):
    rr = proj_r + 1 + yr
    inv.row_dimensions[rr].height = 16
    inv.cell(row=rr, column=2, value=f"Year {yr}").font = ft(size=10)
    inv.cell(row=rr, column=2).border = thin_border("bottom")
    formula_cell(inv, rr, 3, "=C17", color=GREEN)
    inv.cell(row=rr, column=4, value="10%").font = ft(size=10, color="555555")
    inv.cell(row=rr, column=4).alignment = align("center")
    inv.cell(row=rr, column=4).border = thin_border("bottom")
    n = yr * 12
    formula_cell(inv, rr, 5, f"=C{rr}*(((1+10%/12)^{n}-1)/(10%/12))")
    formula_cell(inv, rr, 6, f"=C{rr}*(((1+13%/12)^{n}-1)/(13%/12))")
    for col in [2,3,4,5,6]:
        inv.cell(row=rr, column=col).fill = fill(LGRAY if yr%2==0 else WHITE)

# ═════════════════════════════════════════════════════════════════════════════
# ── SHEET 6: RETIREMENT ───────────────────────────────────────────────────────
ret = wb.create_sheet("🏦 Retirement")
ret.sheet_view.showGridLines = False
set_col_widths(ret, {"A": 3, "B": 34, "C": 20, "D": 20, "E": 3})

ret.row_dimensions[2].height = 32
ret.merge_cells("B2:D2")
th5 = ret["B2"]
th5.value = "🏦  Retirement Corpus Planner"
th5.font = ft(bold=True, size=15, color=WHITE)
th5.fill = fill(NAVY)
th5.alignment = align("center", "center")

section_title(ret, 4, 2, "  ⚙️  RETIREMENT ASSUMPTIONS", 4)
ret_inputs = [
    ("Current Age",                 "='⚙️ Inputs'!C8",  "num",  True),
    ("Target Retirement Age",       60,                  "num",  False),
    ("Years to Retirement",         "=C6-C5",            "num",  True),
    ("Current Monthly Expenses (₹)","='⚙️ Inputs'!C20", "inr",  True),
    ("Inflation Rate (p.a.)",       0.06,                "pct",  False),
    ("Expected Investment Return",  0.12,                "pct",  False),
    ("Safe Withdrawal Rate",        0.04,                "pct",  False),
]
for j, (lbl, val, fmt, is_formula) in enumerate(ret_inputs):
    rr = 5 + j
    ret.row_dimensions[rr].height = 18
    label(ret, rr, 2, lbl, bold=is_formula)
    if is_formula:
        formula_cell(ret, rr, 3, val if isinstance(val, str) else str(val), fmt=fmt, bg=LGRAY)
    else:
        input_cell(ret, rr, 3, val, fmt=fmt)

section_title(ret, 13, 2, "  📊  CORPUS REQUIREMENTS", 4)
label(ret, 14, 2, "Monthly Expense at Retirement (inflation-adj)")
formula_cell(ret, 14, 3, "=C8*(1+C9)^C7", bg=LGRAY)
label(ret, 15, 2, "Annual Expense at Retirement")
formula_cell(ret, 15, 3, "=C14*12", bg=LGRAY)
label(ret, 16, 2, "Required Corpus (25× Rule)")
formula_cell(ret, 16, 3, "=C15/C10", bg=YELLOW)
ret.cell(row=16, column=3).font = ft(bold=True, size=11, color=BLUE_INPUT)

section_title(ret, 18, 2, "  💰  CURRENT TRAJECTORY", 4)
label(ret, 19, 2, "Existing Corpus (all investments today)")
formula_cell(ret, 19, 3, "='⚙️ Inputs'!C31", color=GREEN)
label(ret, 20, 2, "FV of Existing Corpus at Retirement (12%)")
formula_cell(ret, 20, 3, "=C19*(1+C11)^C7", bg=LGRAY)
label(ret, 21, 2, "Additional Corpus Needed")
formula_cell(ret, 21, 3, "=MAX(0,C16-C20)", bg=YELLOW)
ret.cell(row=21, column=3).font = ft(bold=True, size=11, color=RED_BG)

label(ret, 22, 2, "Monthly SIP Needed to Fill Gap")
formula_cell(ret, 22, 3,
    "=IF(C7>0,C21*(C11/12)/((1+C11/12)^(C7*12)-1),0)",
    bg=YELLOW)
ret.cell(row=22, column=3).font = ft(bold=True, size=12, color=BLUE_INPUT)

label(ret, 23, 2, "You are currently investing (monthly)")
formula_cell(ret, 23, 3, "='💼 Investments'!C17", color=GREEN)

label(ret, 24, 2, "Monthly Investment Gap")
formula_cell(ret, 24, 3, "=MAX(0,C22-C23)", bg=LGRAY)
ret.cell(row=24, column=3).font = ft(bold=True, size=11, color=RED_BG)

# Milestones
section_title(ret, 26, 2, "  🗓️  CORPUS MILESTONE PROJECTIONS", 4, bg=TEAL)
header_row(ret, 27, [2,3,4], ["Age","Projected Corpus (at current SIP, 12%)","Target by That Age"], bg=TEAL)
for idx, age_offset in enumerate([5, 10, 15, 20, 25]):
    rr = 28 + idx
    ret.row_dimensions[rr].height = 17
    ret.cell(row=rr, column=2, value=f"=C5+{age_offset}").number_format = "0"
    ret.cell(row=rr, column=2).font = ft(size=10)
    ret.cell(row=rr, column=2).border = thin_border("bottom")
    n = age_offset * 12
    formula_cell(ret, rr, 3,
        f"=C19*(1+C11)^{age_offset}+C23*(((1+C11/12)^{n}-1)/(C11/12))")
    formula_cell(ret, rr, 4,
        f"=C8*(1+C9)^{age_offset}*12*25/C10")
    for col in [2,3,4]:
        ret.cell(row=rr, column=col).fill = fill(LGRAY if idx%2==0 else WHITE)

# ═════════════════════════════════════════════════════════════════════════════
# ── SHEET 7: ESOP PLANNER ────────────────────────────────────────────────────
esop = wb.create_sheet("📊 ESOP Planner")
esop.sheet_view.showGridLines = False
set_col_widths(esop, {"A": 3, "B": 32, "C": 18, "D": 18, "E": 18, "F": 18, "G": 3})

esop.row_dimensions[2].height = 32
esop.merge_cells("B2:F2")
th6 = esop["B2"]
th6.value = "📊  ESOP & Equity Exit Planner (Unlisted Company)"
th6.font = ft(bold=True, size=15, color=WHITE)
th6.fill = fill(NAVY)
th6.alignment = align("center", "center")

section_title(esop, 4, 2, "  ⚙️  YOUR ESOP DETAILS", 6)
esop_inputs = [
    ("ESOPs Granted (shares)",     "='⚙️ Inputs'!C41", "num",  True),
    ("Strike / Exercise Price (₹)","='⚙️ Inputs'!C42", "inr",  True),
    ("FMV per Share Today (₹)",    "='⚙️ Inputs'!C43", "inr",  True),
    ("Your Tax Slab (%)",          0.30,                "pct",  False),
    ("Holding Period after Exercise (months)", 24,      "num",  False),
]
for j, (lbl, val, fmt, is_formula) in enumerate(esop_inputs):
    rr = 5 + j
    esop.row_dimensions[rr].height = 18
    label(esop, rr, 2, lbl, bold=is_formula)
    if is_formula and isinstance(val, str):
        formula_cell(esop, rr, 3, val, fmt=fmt, bg=LGRAY)
    else:
        input_cell(esop, rr, 3, val, fmt=fmt)

# Exercise tax
section_title(esop, 11, 2, "  🧾  TAX AT EXERCISE (Perquisite Tax)", 6, bg=MGRAY)
label(esop, 12, 2, "Perquisite Value per Share (FMV − Strike)")
formula_cell(esop, 12, 3, "=MAX(0,C7-C6)")
label(esop, 13, 2, "Total Perquisite (all shares)")
formula_cell(esop, 13, 3, "=C12*C5")
label(esop, 14, 2, "Tax at Exercise (at your slab rate)")
formula_cell(esop, 14, 3, "=C13*C8", bg=YELLOW)
esop.cell(row=14, column=3).font = ft(bold=True, color=RED_BG)
label(esop, 15, 2, "Net Value after Exercise Tax")
formula_cell(esop, 15, 3, "=C13-C14", bg=LGRAY)

# Exit scenarios
section_title(esop, 17, 2, "  🚀  EXIT SCENARIO MODELLER", 6, bg=TEAL)
header_row(esop, 18, [2,3,4,5,6],
    ["Company Exit Valuation","Your Stake (from Inputs)","Pre-tax Value","Tax (LTCG 20% unlisted)","Net Proceeds"],
    bg=TEAL, size=10)

scenarios = [5_000_000, 25_000_000, 100_000_000, 500_000_000, 1_000_000_000, 0]
stake_ref = "='⚙️ Inputs'!C40"
for j, val in enumerate(scenarios):
    rr = 19 + j
    esop.row_dimensions[rr].height = 18
    esop.cell(row=rr, column=2, value=val).number_format = u'[\u20B9]#,##0;([\u20B9]#,##0);"-"'
    esop.cell(row=rr, column=2).font = ft(size=10, color=BLUE_INPUT)
    esop.cell(row=rr, column=2).fill = fill(YELLOW)
    esop.cell(row=rr, column=2).alignment = align("right")
    esop.cell(row=rr, column=2).border = thin_border()
    formula_cell(esop, rr, 3, stake_ref, fmt="pct", color=GREEN)
    formula_cell(esop, rr, 4, f"=B{rr}*C{rr}")
    formula_cell(esop, rr, 5, f"=D{rr}*0.2", color=RED_BG)
    formula_cell(esop, rr, 6, f"=D{rr}-E{rr}", bg=LGRAY)
    for col in [2,3,4,5,6]:
        esop.cell(row=rr, column=col).fill = fill(LGRAY if j%2==0 else WHITE)
        if col == 2:
            esop.cell(row=rr, column=col).fill = fill(YELLOW)

# Reminder
note_r = 26
esop.merge_cells(f"B{note_r}:F{note_r}")
n_cell = esop.cell(row=note_r, column=2)
n_cell.value = ("⚠️  Note: LTCG on unlisted shares = 20% with indexation (held ≥24 months). "
                "STCG = slab rate (held <24 months). Exercise timing significantly affects tax. Consult your CA.")
n_cell.font = ft(size=9, italic=True, color="555555")
n_cell.alignment = align("left", wrap=True)
esop.row_dimensions[note_r].height = 30

# ═════════════════════════════════════════════════════════════════════════════
# ── SHEET 8: INSURANCE AUDIT ─────────────────────────────────────────────────
ins = wb.create_sheet("🛡️ Insurance Audit")
ins.sheet_view.showGridLines = False
set_col_widths(ins, {"A": 3, "B": 28, "C": 18, "D": 18, "E": 18, "F": 24, "G": 3})

ins.row_dimensions[2].height = 32
ins.merge_cells("B2:F2")
th7 = ins["B2"]
th7.value = "🛡️  Insurance Coverage Audit"
th7.font = ft(bold=True, size=15, color=WHITE)
th7.fill = fill(NAVY)
th7.alignment = align("center", "center")

header_row(ins, 4, [2,3,4,5,6],
    ["Insurance Type","Required Cover","You Have","Gap","Action Required"],
    bg=TEAL)

ins_data = [
    ("Health Insurance",
     "=IF('⚙️ Inputs'!C9=\"Bengaluru\",1500000,IF('⚙️ Inputs'!C9=\"Mumbai\",2000000,1000000))",
     "='⚙️ Inputs'!C34",
     "=MAX(0,C5-D5)",
     "=IF(D5=0,\"🔴 BUY NOW — Niva Bupa / Care Health\",IF(D5<C5,\"🟡 Increase cover\",\"✅ OK\"))"),
    ("Term Life Insurance",
     "=MAX(10000000,'⚙️ Inputs'!C7*12*15)",
     "='⚙️ Inputs'!C35",
     "=MAX(0,C6-D6)",
     "=IF(D6=0,\"🔴 BUY NOW — HDFC Click 2 Protect\",IF(D6<C6,\"🟡 Increase cover\",\"✅ OK\"))"),
    ("Personal Accident",
     5000000,
     0,
     "=MAX(0,C7-D7)",
     "=IF(D7=0,\"🟡 Consider buying — ~₹3,000/yr\",\"✅ OK\")"),
    ("Critical Illness",
     2500000,
     0,
     "=MAX(0,C8-D8)",
     "=IF(D8=0,\"🟡 Consider if 35+ or family history\",\"✅ OK\")"),
]

for j, (ins_type, req, have, gap, action) in enumerate(ins_data):
    rr = 5 + j
    ins.row_dimensions[rr].height = 22
    label(ins, rr, 2, ins_type, bold=True)
    formula_cell(ins, rr, 3, req if isinstance(req, str) else str(req))
    if isinstance(have, str):
        formula_cell(ins, rr, 4, have, color=GREEN)
    else:
        input_cell(ins, rr, 4, have, note="Enter your current cover amount")
    formula_cell(ins, rr, 5, gap, color=RED_BG)
    c_action = ins.cell(row=rr, column=6, value=action)
    c_action.font = ft(size=10, bold=True)
    c_action.alignment = align("left", wrap=True)
    c_action.border = thin_border()
    ins.row_dimensions[rr].height = 25
    for col in [2,3,4,5,6]:
        ins.cell(row=rr, column=col).fill = fill(LGRAY if j%2==0 else WHITE)

# Cost benchmarks
section_title(ins, 10, 2, "  💰  PREMIUM BENCHMARKS (Annual, Online Direct Purchase)", 6, bg=MGRAY)
header_row(ins, 11, [2,3,4,5,6],
    ["Policy Type","Cover Amount","Age Band","Approx Annual Premium","80D Deductible"],
    bg=MGRAY, fg="000000")
benchmarks = [
    ("Health Floater",    "₹10 Lakh",    "30–35",  "₹12,000–18,000",  "Yes (₹25,000 self)"),
    ("Health Floater",    "₹20 Lakh",    "30–35",  "₹18,000–28,000",  "Yes (₹25,000 self)"),
    ("Term Life",         "₹1 Crore",    "28–32",  "₹8,000–12,000",   "Yes (u/s 80C)"),
    ("Term Life",         "₹2 Crore",    "28–32",  "₹14,000–20,000",  "Yes (u/s 80C)"),
    ("Personal Accident", "₹25 Lakh",    "Any",    "₹2,500–4,000",    "No"),
]
for j, row in enumerate(benchmarks):
    rr = 12 + j
    ins.row_dimensions[rr].height = 17
    for col_idx, val in enumerate(row):
        c = ins.cell(row=rr, column=2+col_idx, value=val)
        c.font = ft(size=10)
        c.border = thin_border("bottom")
        c.alignment = align("center" if col_idx > 0 else "left")
        c.fill = fill(LGRAY if j%2==0 else WHITE)

# ═════════════════════════════════════════════════════════════════════════════
# ── SHEET 9: 90-DAY PLAN ─────────────────────────────────────────────────────
plan = wb.create_sheet("✅ 90-Day Plan")
plan.sheet_view.showGridLines = False
set_col_widths(plan, {"A": 3, "B": 6, "C": 34, "D": 22, "E": 14, "F": 16, "G": 3})

plan.row_dimensions[2].height = 32
plan.merge_cells("B2:F2")
th8 = plan["B2"]
th8.value = "✅  90-Day Financial Action Plan"
th8.font = ft(bold=True, size=15, color=WHITE)
th8.fill = fill(NAVY)
th8.alignment = align("center", "center")

header_row(plan, 4, [2,3,4,5,6],
    ["✓","Action","Why It Matters","Deadline","CA/RIA Needed?"],
    bg=TEAL)

actions = [
    # Week 1
    ("WEEK 1 — SURVIVAL FOUNDATION", None, None, None, True),
    ("","Buy personal health insurance (₹10–20L family floater)",
     "Startup group cover doesn't travel with you. One hospitalisation without cover = financial disaster.",
     "Day 3", "No — buy online at Niva Bupa / Care Health"),
    ("","Open a dedicated emergency fund account (liquid MF or separate savings a/c)",
     "Emergency fund must be ring-fenced — never mixed with startup's money or daily expenses.",
     "Day 5", "No — Zerodha Coin or Groww in 10 mins"),
    ("","Calculate exact personal runway",
     "If runway < 6 months, this is your #1 problem — everything else is secondary.",
     "Day 7", "No — use this sheet"),

    # Month 1
    ("MONTH 1 — TAX & INSURANCE", None, None, None, True),
    ("","Check Old vs New Regime for this FY",
     "Wrong regime can cost ₹20,000–50,000+ unnecessarily. Takes 15 minutes with your CA.",
     "Week 2", "Yes — CA or use Tax Planner sheet"),
    ("","Open PPF account and deposit ₹500 to activate",
     "EEE status means 0 tax at all stages. Every month of delay = lost compounding.",
     "Week 2", "No — any bank net banking"),
    ("","Buy term life insurance if you have dependents",
     "₹1 Crore cover costs ~₹10,000/yr at age 30. Your family cannot depend on startup equity.",
     "Week 3", "No — PolicyBazaar comparison"),
    ("","Mark advance tax dates in calendar (15 Jun/Sep/Dec/Mar)",
     "234B interest penalty if underpaid. Easy to miss as a self-employed founder.",
     "Week 4", "Yes — CA for calculation"),

    # Month 2
    ("MONTH 2 — RETIREMENT FOUNDATIONS", None, None, None, True),
    ("","Open NPS Tier 1 on eNPS.nsdl.com",
     "₹50,000/yr = ₹15,600 tax saved (30% bracket). Extra over 80C limit. One-time setup.",
     "Week 5", "No — eNPS.nsdl.com"),
    ("","Start ELSS SIP for remaining 80C room",
     "3-year lock-in is shortest among 80C options. Market-linked returns ~12-14% historically.",
     "Week 5", "No — Zerodha Coin, direct plan"),
    ("","Start Nifty 50 Index Fund SIP (even ₹1,000/month)",
     "Low cost (0.1%), passive, no fund manager risk. Time in market beats timing the market.",
     "Week 6", "No — Zerodha Coin or AMC website"),
    ("","Withdraw and reinvest EPF from old employer (if applicable)",
     "Money sitting in an old EPF account earns ~8.15% — comparable but less flexible than NPS.",
     "Week 7", "Yes — CA for transfer advice"),

    # Month 3
    ("MONTH 3 — OPTIMISE & REVIEW", None, None, None, True),
    ("","Review SIP amounts — can you increase?",
     "Every ₹1,000/month extra SIP at 12% = ₹19.7L over 20 years. Small increases compound hugely.",
     "Week 9", "No"),
    ("","File advance tax if tax liability > ₹10,000",
     "15 June deadline. If salary TDS covers your liability, this may not be needed — verify.",
     "Before Jun 15","Yes — CA"),
    ("","Consult a SEBI-registered RIA for a full review",
     "One session (~₹2,000-5,000) with a fee-only RIA gives personalised, unbiased advice.",
     "Week 12", "Yes — find at sebi.gov.in"),
    ("","Update this spreadsheet with actual numbers",
     "A plan only works if the numbers are current. Review monthly.",
     "Month-end", "No"),
]

r = 5
for item in actions:
    plan.row_dimensions[r].height = 20 if not item[4] else 22
    if item[4]:  # section header
        section_title(plan, r, 2, f"  {item[0]}", 6, bg=TEAL if "WEEK" in item[0] else ORANGE)
        plan.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    else:
        chk = plan.cell(row=r, column=2, value="☐")
        chk.font = ft(size=13, color=ORANGE)
        chk.alignment = align("center", "center")
        chk.border = thin_border("bottom")

        act = plan.cell(row=r, column=3, value=item[1])
        act.font = ft(size=10)
        act.alignment = align("left", wrap=True)
        act.border = thin_border("bottom")

        why = plan.cell(row=r, column=4, value=item[2])
        why.font = ft(size=9, italic=True, color="444444")
        why.alignment = align("left", wrap=True)
        why.border = thin_border("bottom")

        dl = plan.cell(row=r, column=5, value=item[3])
        dl.font = ft(size=10, bold=True, color=RED_BG)
        dl.alignment = align("center")
        dl.border = thin_border("bottom")

        ca = plan.cell(row=r, column=6, value=item[4] if isinstance(item[4], str) else "")
        ca.font = ft(size=9, color="555555")
        ca.alignment = align("left", wrap=True)
        ca.border = thin_border("bottom")

        bg_col = LGRAY if r % 2 == 0 else WHITE
        for col in [2,3,4,5,6]:
            plan.cell(row=r, column=col).fill = fill(bg_col)
    r += 1

# ── Tab colors ────────────────────────────────────────────────────────────────
dash.sheet_properties.tabColor  = "1B3A6B"
inp.sheet_properties.tabColor   = "E07B39"
cf.sheet_properties.tabColor    = "0E6B6B"
tax.sheet_properties.tabColor   = "C95E1E"
inv.sheet_properties.tabColor   = "1A6B3C"
ret.sheet_properties.tabColor   = "2E86AB"
esop.sheet_properties.tabColor  = "8E44AD"
ins.sheet_properties.tabColor   = "C0392B"
plan.sheet_properties.tabColor  = "1A6B3C"

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(OUTPUT)
print(f"✅ Saved: {OUTPUT}")
